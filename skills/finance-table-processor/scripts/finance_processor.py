#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
财务表格处理器 - 专门处理收支明细表
功能：读取、分析、汇总、生成报表
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import sys
import argparse
import os

class FinanceTableProcessor:
    """财务表格处理器"""
    
    def __init__(self, file_path):
        self.file_path = file_path
        self.df = None
        self.summary = {}
        
    def load_data(self):
        """加载数据"""
        ext = os.path.splitext(self.file_path)[1].lower()
        
        try:
            if ext in ['.xlsx', '.xls']:
                self.df = pd.read_excel(self.file_path)
            elif ext == '.csv':
                self.df = pd.read_csv(self.file_path)
            else:
                raise ValueError(f"不支持的文件格式: {ext}")
            
            print(f"✅ 成功加载数据: {len(self.df)} 行, {len(self.df.columns)} 列")
            return True
        except Exception as e:
            print(f"❌ 加载失败: {e}")
            return False
    
    def analyze_structure(self):
        """分析表格结构"""
        print("\n📊 表格结构分析:")
        print(f"列名: {list(self.df.columns)}")
        print(f"数据类型:\n{self.df.dtypes}")
        print(f"\n前5行预览:\n{self.df.head()}")
        
    def auto_detect_columns(self):
        """自动识别关键列"""
        columns = {col.lower(): col for col in self.df.columns}
        
        # 日期列
        date_cols = ['date', '日期', '时间', 'time', 'datetime']
        self.date_col = next((columns[c] for c in date_cols if c in columns), None)
        
        # 金额列
        amount_cols = ['amount', '金额', 'money', 'price', 'sum', 'total']
        self.amount_col = next((columns[c] for c in amount_cols if c in columns), None)
        
        # 类型/分类列
        type_cols = ['type', '类型', 'category', '分类', '收支', 'income_expense']
        self.type_col = next((columns[c] for c in type_cols if c in columns), None)
        
        # 备注/描述列
        desc_cols = ['description', '备注', 'desc', '说明', '摘要', 'content']
        self.desc_col = next((columns[c] for c in desc_cols if c in columns), None)
        
        print(f"\n🔍 自动识别列:")
        print(f"  日期列: {self.date_col}")
        print(f"  金额列: {self.amount_col}")
        print(f"  类型列: {self.type_col}")
        print(f"  备注列: {self.desc_col}")
        
    def clean_data(self):
        """数据清洗"""
        if self.date_col:
            # 转换日期格式
            self.df[self.date_col] = pd.to_datetime(self.df[self.date_col], errors='coerce')
            # 添加年月列用于分组
            self.df['年月'] = self.df[self.date_col].dt.to_period('M').astype(str)
        
        if self.amount_col:
            # 清理金额（去除货币符号、逗号）
            self.df[self.amount_col] = pd.to_numeric(
                self.df[self.amount_col].astype(str).str.replace('[￥$,]', '', regex=True),
                errors='coerce'
            )
        
        # 删除空行
        self.df = self.df.dropna(subset=[self.amount_col] if self.amount_col else None, how='all')
        
        print(f"\n🧹 数据清洗完成: {len(self.df)} 行有效数据")
        
    def classify_income_expense(self):
        """自动分类收入和支出"""
        if not self.amount_col:
            return
        
        # 如果没有类型列，根据金额正负自动分类
        if not self.type_col:
            self.df['收支类型'] = self.df[self.amount_col].apply(
                lambda x: '收入' if x > 0 else '支出' if x < 0 else '其他'
            )
            self.type_col = '收支类型'
        
        # 确保金额为正数（支出用负数表示）
        self.df['金额_正'] = self.df[self.amount_col].abs()
        
    def calculate_summary(self):
        """计算汇总统计"""
        if not self.amount_col:
            print("❌ 未找到金额列，无法计算汇总")
            return
        
        # 总体统计
        total_income = self.df[self.df[self.type_col] == '收入'][self.amount_col].sum() if self.type_col else 0
        total_expense = self.df[self.df[self.type_col] == '支出'][self.amount_col].sum() if self.type_col else 0
        
        self.summary = {
            '总记录数': len(self.df),
            '总收入': total_income,
            '总支出': abs(total_expense),
            '净收支': total_income + total_expense,
            '平均收入': self.df[self.df[self.type_col] == '收入'][self.amount_col].mean() if self.type_col and total_income > 0 else 0,
            '平均支出': self.df[self.df[self.type_col] == '支出'][self.amount_col].mean() if self.type_col and total_expense < 0 else 0,
        }
        
        print("\n📈 汇总统计:")
        for key, value in self.summary.items():
            if isinstance(value, float):
                print(f"  {key}: ¥{value:,.2f}")
            else:
                print(f"  {key}: {value}")
                
    def monthly_report(self):
        """生成月度报表"""
        if '年月' not in self.df.columns or not self.type_col:
            print("❌ 缺少必要的列，无法生成月度报表")
            return None
        
        monthly = self.df.groupby(['年月', self.type_col])[self.amount_col].sum().unstack(fill_value=0)
        monthly['净收支'] = monthly.get('收入', 0) + monthly.get('支出', 0)
        
        print("\n📅 月度报表:")
        print(monthly.to_string())
        
        return monthly
    
    def export_report(self, output_path):
        """导出报表到Excel"""
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # 原始数据
            self.df.to_excel(writer, sheet_name='原始数据', index=False)
            
            # 汇总统计
            summary_df = pd.DataFrame([self.summary])
            summary_df.to_excel(writer, sheet_name='汇总统计', index=False)
            
            # 月度报表
            monthly = self.monthly_report()
            if monthly is not None:
                monthly.to_excel(writer, sheet_name='月度报表')
        
        # 美化格式
        self._format_excel(output_path)
        print(f"\n✅ 报表已导出: {output_path}")
        
    def _format_excel(self, file_path):
        """美化Excel格式"""
        wb = openpyxl.load_workbook(file_path)
        
        # 定义样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        money_font = Font(color="FF0000")  # 支出红色
        income_font = Font(color="00B050")  # 收入绿色
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 设置列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 表头样式
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
        
        wb.save(file_path)

def main():
    parser = argparse.ArgumentParser(description='财务表格处理器')
    parser.add_argument('file', help='输入文件路径 (.xlsx, .xls, .csv)')
    parser.add_argument('-o', '--output', help='输出报表路径')
    parser.add_argument('--analyze', action='store_true', help='仅分析结构')
    
    args = parser.parse_args()
    
    # 初始化处理器
    processor = FinanceTableProcessor(args.file)
    
    # 加载数据
    if not processor.load_data():
        sys.exit(1)
    
    # 分析结构
    if args.analyze:
        processor.analyze_structure()
        return
    
    # 完整处理流程
    processor.analyze_structure()
    processor.auto_detect_columns()
    processor.clean_data()
    processor.classify_income_expense()
    processor.calculate_summary()
    processor.monthly_report()
    
    # 导出报表
    if args.output:
        processor.export_report(args.output)
    else:
        default_output = f"财务报表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        processor.export_report(default_output)

if __name__ == '__main__':
    main()
